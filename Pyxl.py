import openpyxl

wb = openpyxl.load_workbook('transactions.xlsx')
print(wb.sheetnames)

sheet = wb['Sheet1']
cell = sheet['a1']
print(cell)
print(cell.value)
print(cell.row)
print(sheet.cell(1, 1))
print(sheet.max_row)
print(sheet.max_column)

for row in range(1, sheet.max_row + 1):
    for column in range(1, sheet.max_column + 1):
        cell = sheet.cell(row, column)
        print(cell.value)

# sheet.append([4, 'Parth', 100000])
